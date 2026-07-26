from __future__ import annotations

import json
import math
import textwrap
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "dashboard-import-workbook"
OUTPUT_FILE = OUTPUT_DIR / "Post E Dashboard Complete Import Template.xlsx"
PREVIEW_DIR = OUTPUT_DIR / "previews"

NAVY = "172033"
NAVY_2 = "26344B"
BLUE = "6E8FBE"
PALE_BLUE = "EAF1FB"
CREAM = "F7F2ED"
WHITE = "FFFFFF"
INK = "223047"
MUTED = "65738A"
GRID = "D9E2EF"
GREEN = "DFF2E5"
AMBER = "FFF0CC"
RED = "F9DFDF"
PALE_PURPLE = "EEE9FA"

THIN_BORDER = Border(
    left=Side(style="thin", color=GRID),
    right=Side(style="thin", color=GRID),
    top=Side(style="thin", color=GRID),
    bottom=Side(style="thin", color=GRID),
)


def add_header_comments(ws, comments: dict[str, str]) -> None:
    for cell in ws[1]:
        if cell.value in comments:
            cell.comment = Comment(comments[cell.value], "Post E Dashboard")


def style_data_sheet(
    ws,
    widths: dict[str, float],
    comments: dict[str, str],
    table_name: str,
    tab_color: str,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = tab_color
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddHeader.center.text = f"&B{ws.title}"
    ws.oddFooter.right.text = "Page &P of &N"

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos Display", size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=CREAM if cell.row % 2 == 0 else WHITE)
            cell.font = Font(name="Aptos", size=10, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

    add_header_comments(ws, comments)

    ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def list_validation(values: list[str], prompt: str) -> DataValidation:
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = f"Choose one of: {', '.join(values)}"
    validation.errorTitle = "Invalid selection"
    validation.prompt = prompt
    validation.promptTitle = "Dashboard input"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    return validation


def whole_validation(prompt: str) -> DataValidation:
    validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="1000000000",
        allow_blank=True,
    )
    validation.error = "Enter a whole number of zero or more."
    validation.errorTitle = "Invalid number"
    validation.prompt = prompt
    validation.promptTitle = "Dashboard input"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    return validation


def decimal_validation(prompt: str) -> DataValidation:
    validation = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1000000000",
        allow_blank=True,
    )
    validation.error = "Enter a number of zero or more."
    validation.errorTitle = "Invalid number"
    validation.prompt = prompt
    validation.promptTitle = "Dashboard input"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    return validation


def date_validation(prompt: str) -> DataValidation:
    validation = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2020,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    validation.error = "Enter a valid date between 2020 and 2100."
    validation.errorTitle = "Invalid date"
    validation.prompt = prompt
    validation.promptTitle = "Dashboard input"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    return validation


def build_guide(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "START HERE"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY
    ws.freeze_panes = "A10"
    ws.merge_cells("A1:H1")
    ws["A1"] = "Post E Dashboard — Complete Import Workbook"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 44

    ws.merge_cells("A2:H2")
    ws["A2"] = "Replace the sample rows with your data, keep every sheet name and header unchanged, save as .xlsx, then import the whole workbook."
    ws["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws["A2"].font = Font(name="Aptos", size=11, italic=True, color=INK)
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    cards = [
        ("A4:B6", "1  Replace examples", "Edit data from row 2 downward on each import sheet."),
        ("C4:D6", "2  Keep structure", "Do not rename tabs or row-1 headers. Do not add title rows."),
        ("E4:F6", "3  Reconcile totals", "Issue counts should match Total Threads; owner totals should match current on-hold."),
        ("G4:H6", "4  Import workbook", "Save as .xlsx and use the dashboard import control."),
    ]
    for cell_range, title, body in cards:
        ws.merge_cells(cell_range)
        top_left = ws[cell_range.split(":")[0]]
        top_left.value = f"{title}\n{body}"
        top_left.fill = PatternFill("solid", fgColor=CREAM)
        top_left.font = Font(name="Aptos", size=11, bold=True, color=INK)
        top_left.alignment = Alignment(vertical="center", wrap_text=True)
        top_left.border = Border(
            left=Side(style="medium", color=BLUE),
            right=Side(style="thin", color=GRID),
            top=Side(style="thin", color=GRID),
            bottom=Side(style="thin", color=GRID),
        )
    for row in range(4, 7):
        ws.row_dimensions[row].height = 28

    headers = ["Sheet", "Required row-1 headers", "Dashboard area filled", "Important rule", "Example rows"]
    start_row = 10
    for col, value in enumerate(headers, 1):
        cell = ws.cell(start_row, col, value)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    guide_rows = [
        (
            "Dashboard Summary",
            "WW | Total WIP | Total Threads | Completion % | Current/Previous Onhold Lots | Current/Previous Onhold WIP",
            "Work week, KPI cards, completion, previous on-hold comparison",
            "Use 0–100 for Completion %. Detail sheets may recalculate current totals.",
            "1",
        ),
        (
            "Issue Contributors",
            "WW | Icon | Issue | Tool | Count | Shift A | Shift B | Shift C | Shift D",
            "Trend chart, top contributors, insights, total threads",
            "Count should equal Shift A+B+C+D for each row.",
            "4",
        ),
        (
            "Summary Cases",
            "Icon | Category | Status | Count | Item | Image",
            "Weekly summary cards and case details",
            "Use one row per detail; repeat the category and category total.",
            "5",
        ),
        (
            "Onhold Owners",
            "Owner | Lots | WIP",
            "Owner breakdown and current on-hold totals",
            "Current totals are the sum of all owner rows.",
            "4",
        ),
        (
            "Reminders",
            "Title | Message",
            "Sidebar reminder carousel",
            "Use Alt+Enter for multiple lines inside Message.",
            "2",
        ),
        (
            "Calendar Events",
            "Date | Title | Detail",
            "Upcoming calendar list and chatbot calendar answer",
            "Use real Excel dates. Add only extra events; public holidays are built in.",
            "2",
        ),
        (
            "Feedback Evidence",
            "Issue | Triggered Tool | Shift | Date | Notes | Images",
            "Feedback evidence and tool-trigger chart",
            "Separate multiple image URLs/file names with commas, |, or line breaks.",
            "2",
        ),
    ]
    for row_idx, row in enumerate(guide_rows, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.fill = PatternFill("solid", fgColor=WHITE if row_idx % 2 else CREAM)
            cell.font = Font(name="Aptos", size=10, color=INK, bold=col_idx == 1)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[row_idx].height = 54

    ws.merge_cells("A20:H20")
    ws["A20"] = (
        "Import precedence: Issue Contributors updates Total Threads; Onhold Owners updates current on-hold Lots/WIP. "
        "Keep the summary values aligned with those detail-sheet totals."
    )
    ws["A20"].fill = PatternFill("solid", fgColor=AMBER)
    ws["A20"].font = Font(name="Aptos", size=10, bold=True, color=INK)
    ws["A20"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[20].height = 34

    for column, width in {
        "A": 22,
        "B": 31,
        "C": 23,
        "D": 31,
        "E": 12,
        "F": 15,
        "G": 15,
        "H": 15,
    }.items():
        ws.column_dimensions[column].width = width
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard Summary")
    ws.append(
        [
            "WW",
            "Total WIP",
            "Total Threads",
            "Completion %",
            "Current Onhold Lots",
            "Previous Onhold Lots",
            "Current Onhold WIP",
            "Previous Onhold WIP",
        ]
    )
    ws.append(["WW21/22", 15.2, 424, 78, 179, 190, 30, 34])
    style_data_sheet(
        ws,
        {"A": 14, "B": 14, "C": 16, "D": 16, "E": 22, "F": 23, "G": 22, "H": 23},
        {
            "WW": "Dashboard work-week label, for example WW21/22.",
            "Total WIP": "Post E total WIP. Enter a number, not text.",
            "Total Threads": "Total Microsoft Teams threads. Keep aligned with the sum of Issue Contributors Count.",
            "Completion %": "Whole percentage from 0 to 100. Example: enter 78, not 0.78.",
            "Current Onhold Lots": "Current on-hold lot count. Keep aligned with Onhold Owners total Lots.",
            "Previous Onhold Lots": "Previous-week on-hold lot count.",
            "Current Onhold WIP": "Current on-hold WIP. Keep aligned with Onhold Owners total WIP.",
            "Previous Onhold WIP": "Previous-week on-hold WIP.",
        },
        "tblDashboardSummary",
        BLUE,
    )
    ws["B2"].number_format = "0.0"
    ws["C2"].number_format = "0"
    ws["D2"].number_format = '0"%"'
    for cell in ["E2", "F2"]:
        ws[cell].number_format = "0"
    for cell in ["G2", "H2"]:
        ws[cell].number_format = "0.0"

    dv_wip = decimal_validation("Enter a WIP value of zero or more.")
    ws.add_data_validation(dv_wip)
    dv_wip.add("B2:B500")
    dv_count = whole_validation("Enter a whole-number count of zero or more.")
    ws.add_data_validation(dv_count)
    dv_count.add("C2:C500")
    dv_completion = DataValidation(
        type="whole", operator="between", formula1="0", formula2="100", allow_blank=True
    )
    dv_completion.error = "Enter a whole percentage from 0 to 100."
    dv_completion.errorTitle = "Invalid completion"
    dv_completion.prompt = "Enter 78 for 78%, not 0.78."
    dv_completion.promptTitle = "Completion %"
    dv_completion.showErrorMessage = True
    dv_completion.showInputMessage = True
    ws.add_data_validation(dv_completion)
    dv_completion.add("D2:D500")
    dv_lots = whole_validation("Enter a whole-number lot count of zero or more.")
    ws.add_data_validation(dv_lots)
    dv_lots.add("E2:F500")
    dv_onhold_wip = decimal_validation("Enter an on-hold WIP value of zero or more.")
    ws.add_data_validation(dv_onhold_wip)
    dv_onhold_wip.add("G2:H500")


def build_issues(wb: Workbook) -> None:
    ws = wb.create_sheet("Issue Contributors")
    ws.append(["WW", "Icon", "Issue", "Tool", "Count", "Shift A", "Shift B", "Shift C", "Shift D"])
    rows = [
        ["WW21/22", "X", "Mark Invalid", "KLAT790-0026", 239, 55, 60, 56, 68],
        ["WW21/22", "DM", "Data Matrix Invalid", "KLAT790-0036", 64, 30, 21, 10, 3],
        ["WW21/22", "SN", "SafetyNet Issue", "KLAT790-0090", 36, 15, 15, 3, 3],
        ["WW21/22", "OT", "Others", "KLAT790-0090", 85, 17, 25, 29, 14],
    ]
    for row in rows:
        ws.append(row)
    style_data_sheet(
        ws,
        {"A": 13, "B": 10, "C": 25, "D": 20, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12},
        {
            "WW": "Work week for contributor filtering, for example WW21/22.",
            "Icon": "Short label or emoji displayed with the contributor.",
            "Issue": "Issue/category name used on the trend chart.",
            "Tool": "Machine or triggered tool identifier.",
            "Count": "Total mentions for this row. Keep equal to Shift A+B+C+D.",
            "Shift A": "Mentions attributed to Shift A.",
            "Shift B": "Mentions attributed to Shift B.",
            "Shift C": "Mentions attributed to Shift C.",
            "Shift D": "Mentions attributed to Shift D.",
        },
        "tblIssueContributors",
        BLUE,
    )
    for column in "EFGHI":
        for cell in ws[column][1:]:
            cell.number_format = "0"
    dv = whole_validation("Enter a whole-number count of zero or more.")
    ws.add_data_validation(dv)
    dv.add("E2:I500")


def build_cases(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary Cases")
    ws.append(["Icon", "Category", "Status", "Count", "Item", "Image"])
    rows = [
        ["MRB", "MRB", "Ongoing", 2, "MRB | Post E | Dropped Lots at Cargo Lift M4 and discuss lot disposition.", ""],
        ["MRB", "MRB", "Ongoing", 2, "MRB | Post E | Laser Marking | JDZN3HJ.11 Mark Off Center.", ""],
        ["AQL", "AQL Finding", "Pending", 1, "Additional lot with marker marking on unit. Pending QA meeting with REL team.", ""],
        ["SN", "SafetyNet Issue", "Closed", 2, "Few lots triggered no SafetyNet for NC23M Series. Proceed release with bypass.", ""],
        ["SN", "SafetyNet Issue", "Closed", 2, "2DID shortage from Assembly MMP. DONE MRB with QA and Assembly.", ""],
    ]
    for row in rows:
        ws.append(row)
    style_data_sheet(
        ws,
        {"A": 10, "B": 24, "C": 14, "D": 12, "E": 70, "F": 38},
        {
            "Icon": "Short label or emoji for the case category.",
            "Category": "Rows with the same category are grouped into one dashboard card.",
            "Status": "Choose Pending, Ongoing, or Closed.",
            "Count": "Total number of cases in this category. Repeat the same category total on its detail rows.",
            "Item": "One case detail per row.",
            "Image": "Optional image URL, file name, or data URI associated with the case.",
        },
        "tblSummaryCases",
        PALE_PURPLE,
    )
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 46
        ws[f"D{row}"].number_format = "0"
    dv_status = list_validation(["Pending", "Ongoing", "Closed"], "Select the case status.")
    ws.add_data_validation(dv_status)
    dv_status.add("C2:C500")
    dv_count = whole_validation("Enter the category total as a whole number.")
    ws.add_data_validation(dv_count)
    dv_count.add("D2:D500")
    ws.conditional_formatting.add(
        "C2:C500",
        FormulaRule(formula=['$C2="Closed"'], fill=PatternFill("solid", fgColor=GREEN)),
    )
    ws.conditional_formatting.add(
        "C2:C500",
        FormulaRule(formula=['$C2="Pending"'], fill=PatternFill("solid", fgColor=AMBER)),
    )
    ws.conditional_formatting.add(
        "C2:C500",
        FormulaRule(formula=['$C2="Ongoing"'], fill=PatternFill("solid", fgColor=PALE_BLUE)),
    )


def build_owners(wb: Workbook) -> None:
    ws = wb.create_sheet("Onhold Owners")
    ws.append(["Owner", "Lots", "WIP"])
    for row in [["QA", 68, 11], ["Process", 54, 9], ["Equipment", 32, 6], ["Production", 25, 4]]:
        ws.append(row)
    style_data_sheet(
        ws,
        {"A": 26, "B": 16, "C": 16},
        {
            "Owner": "Responsible on-hold owner/team.",
            "Lots": "Current on-hold lot count for this owner.",
            "WIP": "Current on-hold WIP for this owner.",
        },
        "tblOnholdOwners",
        BLUE,
    )
    dv_lots = whole_validation("Enter a whole-number lot count.")
    ws.add_data_validation(dv_lots)
    dv_lots.add("B2:B500")
    dv_wip = decimal_validation("Enter the owner's current on-hold WIP.")
    ws.add_data_validation(dv_wip)
    dv_wip.add("C2:C500")
    for cell in ws["B"][1:]:
        cell.number_format = "0"
    for cell in ws["C"][1:]:
        cell.number_format = "0.0"


def build_reminders(wb: Workbook) -> None:
    ws = wb.create_sheet("Reminders")
    ws.append(["Title", "Message"])
    ws.append(
        [
            "CAT Test Building Team Q2",
            "Shift C & D -> 25 June 2026\nShift A & B -> 29 June 2026\nOffice Hours & Permanent Day -> To be confirmed by Supervisors",
        ]
    )
    ws.append(["Family Day Registration is NOW OPEN!", "Date: 20/6/2026\nVenue: Micron"])
    style_data_sheet(
        ws,
        {"A": 38, "B": 90},
        {
            "Title": "Reminder headline displayed in the sidebar.",
            "Message": "Reminder details. Use Alt+Enter to add multiple lines inside one cell.",
        },
        "tblReminders",
        AMBER,
    )
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 64


def build_calendar(wb: Workbook) -> None:
    ws = wb.create_sheet("Calendar Events")
    ws.append(["Date", "Title", "Detail"])
    ws.append(
        [
            datetime(2026, 6, 25),
            "CAT Test Building Team Q2 - Shift C & D",
            "Public holidays are built into the dashboard. Add only extra dashboard events here.",
        ]
    )
    ws.append(
        [
            datetime(2026, 6, 29),
            "CAT Test Building Team Q2 - Shift A & B",
            "Office hours and permanent day to be confirmed by supervisors.",
        ]
    )
    style_data_sheet(
        ws,
        {"A": 16, "B": 48, "C": 76},
        {
            "Date": "Real Excel event date. Recommended display: yyyy-mm-dd.",
            "Title": "Event title shown in the calendar/upcoming list.",
            "Detail": "Optional event detail shown by the dashboard and chatbot.",
        },
        "tblCalendarEvents",
        GREEN,
    )
    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 42
    dv = date_validation("Enter a real Excel date.")
    ws.add_data_validation(dv)
    dv.add("A2:A500")


def build_feedback(wb: Workbook) -> None:
    ws = wb.create_sheet("Feedback Evidence")
    ws.append(["Issue", "Triggered Tool", "Shift", "Date", "Notes", "Images"])
    ws.append(
        [
            "Mark Invalid",
            "KLAT790-0035",
            "Shift A",
            datetime(2026, 6, 21),
            "Tool trigger evidence logged from dashboard. Replace this example with the actual evidence note.",
            "",
        ]
    )
    ws.append(
        [
            "Data Matrix Invalid",
            "KLAT790-0036",
            "Shift B",
            datetime(2026, 6, 21),
            "Example feedback evidence row. This sheet also feeds the Tool Trigger Frequency chart.",
            "",
        ]
    )
    style_data_sheet(
        ws,
        {"A": 28, "B": 22, "C": 14, "D": 16, "E": 70, "F": 50},
        {
            "Issue": "Issue title for this evidence record.",
            "Triggered Tool": "Machine/tool identifier used in the trigger-frequency chart.",
            "Shift": "Choose Shift A, Shift B, Shift C, or Shift D.",
            "Date": "Real Excel date for the evidence record.",
            "Notes": "Improvement, disposition, or investigation notes.",
            "Images": "Optional image URLs/file names. Separate multiple values with commas, |, or line breaks.",
        },
        "tblFeedbackEvidence",
        PALE_PURPLE,
    )
    for cell in ws["D"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 50
    dv_shift = list_validation(["Shift A", "Shift B", "Shift C", "Shift D"], "Select the shift.")
    ws.add_data_validation(dv_shift)
    dv_shift.add("C2:C500")
    dv_date = date_validation("Enter a real Excel date.")
    ws.add_data_validation(dv_date)
    dv_date.add("D2:D500")


def build_workbook() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.properties.creator = "OpenAI Codex"
    wb.properties.title = "Post E Dashboard Complete Import Template"
    wb.properties.subject = "Import-ready workbook for Post E Dashboard"
    wb.properties.description = (
        "One workbook containing every data sheet supported by the Post E Dashboard importer."
    )
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    build_guide(wb)
    build_summary(wb)
    build_issues(wb)
    build_cases(wb)
    build_owners(wb)
    build_reminders(wb)
    build_calendar(wb)
    build_feedback(wb)

    wb.active = 0
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


def font(size: int, bold: bool = False):
    preferred = [
        Path("C:/Windows/Fonts/aptos.ttf"),
        Path("C:/Windows/Fonts/arial.ttf"),
    ]
    preferred_bold = [
        Path("C:/Windows/Fonts/aptos-bold.ttf"),
        Path("C:/Windows/Fonts/arialbd.ttf"),
    ]
    for candidate in preferred_bold if bold else preferred:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default()


def cell_fill_hex(cell) -> str:
    color = cell.fill.fgColor
    if cell.fill.fill_type == "solid" and color.type == "rgb" and color.rgb:
        return color.rgb[-6:]
    return WHITE


def render_sheet_preview(ws, path: Path) -> None:
    max_row = min(ws.max_row, 24)
    max_col = min(ws.max_column, 10)
    col_widths = []
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        width = ws.column_dimensions[letter].width or 12
        col_widths.append(max(70, min(330, int(width * 7.2))))
    row_heights = []
    for row_idx in range(1, max_row + 1):
        height = ws.row_dimensions[row_idx].height or 24
        row_heights.append(max(28, min(110, int(height * 1.35))))

    margin = 18
    image_width = sum(col_widths) + margin * 2
    image_height = sum(row_heights) + margin * 2 + 34
    canvas = Image.new("RGB", (image_width, image_height), "#EEF3F9")
    draw = ImageDraw.Draw(canvas)
    draw.rounded_rectangle(
        (8, 8, image_width - 8, image_height - 8),
        radius=14,
        fill="#FFFFFF",
        outline="#CCD8E7",
        width=2,
    )
    draw.text((margin, 16), ws.title, fill=f"#{NAVY}", font=font(18, bold=True))

    y = margin + 34
    merged_map = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col_m, max_row_m = merged.bounds
        merged_map[(min_row, min_col)] = (max_row_m, max_col_m)

    for row_idx in range(1, max_row + 1):
        x = margin
        col_idx = 1
        while col_idx <= max_col:
            cell = ws.cell(row_idx, col_idx)
            merged = merged_map.get((row_idx, col_idx))
            span_col = min(merged[1], max_col) if merged else col_idx
            width = sum(col_widths[col_idx - 1 : span_col])
            height = row_heights[row_idx - 1]
            fill = f"#{cell_fill_hex(cell)}"
            draw.rectangle((x, y, x + width, y + height), fill=fill, outline=f"#{GRID}", width=1)

            value = "" if cell.value is None else str(cell.value)
            if isinstance(cell.value, datetime):
                value = cell.value.strftime("%Y-%m-%d")
            size = 12 if row_idx == 1 else 11
            is_bold = bool(cell.font.bold) or row_idx == 1
            text_color = cell.font.color
            if text_color and text_color.type == "rgb" and text_color.rgb:
                color = f"#{text_color.rgb[-6:]}"
            else:
                color = f"#{INK}"
            max_chars = max(8, int(width / (size * 0.62)))
            wrapped = textwrap.wrap(value, width=max_chars, replace_whitespace=False)
            max_lines = max(1, int((height - 10) / (size + 3)))
            if len(wrapped) > max_lines:
                wrapped = wrapped[:max_lines]
                wrapped[-1] = wrapped[-1][: max(1, max_chars - 1)] + "…"
            draw.multiline_text(
                (x + 7, y + 6),
                "\n".join(wrapped),
                fill=color,
                font=font(size, bold=is_bold),
                spacing=3,
            )
            x += width
            col_idx = span_col + 1
        y += row_heights[row_idx - 1]

    if image_width > 1800:
        scale = 1800 / image_width
        canvas = canvas.resize((1800, math.ceil(image_height * scale)), Image.Resampling.LANCZOS)
    canvas.save(path)


def render_all_sheets(path: Path) -> list[Path]:
    wb = load_workbook(path, data_only=False)
    preview_paths = []
    for index, ws in enumerate(wb.worksheets, 1):
        safe_name = "".join(c if c.isalnum() else "_" for c in ws.title).strip("_")
        preview_path = PREVIEW_DIR / f"{index:02d}_{safe_name}.png"
        render_sheet_preview(ws, preview_path)
        preview_paths.append(preview_path)

    thumbs = []
    for preview_path in preview_paths:
        img = Image.open(preview_path).convert("RGB")
        img.thumbnail((780, 430))
        thumbs.append(img.copy())
    cols = 2
    rows = math.ceil(len(thumbs) / cols)
    contact = Image.new("RGB", (1600, rows * 455), "#DDE6F1")
    for idx, img in enumerate(thumbs):
        x = (idx % cols) * 800 + 10
        y = (idx // cols) * 455 + 10
        contact.paste(img, (x, y))
    contact_path = PREVIEW_DIR / "00_all_sheets_contact.png"
    contact.save(contact_path)
    return [contact_path, *preview_paths]


def validate_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    required_headers = {
        "Dashboard Summary": [
            "WW",
            "Total WIP",
            "Total Threads",
            "Completion %",
            "Current Onhold Lots",
            "Previous Onhold Lots",
            "Current Onhold WIP",
            "Previous Onhold WIP",
        ],
        "Issue Contributors": [
            "WW",
            "Icon",
            "Issue",
            "Tool",
            "Count",
            "Shift A",
            "Shift B",
            "Shift C",
            "Shift D",
        ],
        "Summary Cases": ["Icon", "Category", "Status", "Count", "Item", "Image"],
        "Onhold Owners": ["Owner", "Lots", "WIP"],
        "Reminders": ["Title", "Message"],
        "Calendar Events": ["Date", "Title", "Detail"],
        "Feedback Evidence": ["Issue", "Triggered Tool", "Shift", "Date", "Notes", "Images"],
    }
    errors = []
    for sheet, headers in required_headers.items():
        if sheet not in wb.sheetnames:
            errors.append(f"Missing sheet: {sheet}")
            continue
        ws = wb[sheet]
        actual = [ws.cell(1, i).value for i in range(1, len(headers) + 1)]
        if actual != headers:
            errors.append(f"{sheet} headers differ: {actual}")

    formulas = []
    formula_errors = []
    excel_error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{ws.title}!{cell.coordinate}")
                if isinstance(cell.value, str) and any(token in cell.value for token in excel_error_tokens):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}: {cell.value}")

    summary = wb["Dashboard Summary"]
    issues = wb["Issue Contributors"]
    owners = wb["Onhold Owners"]
    expected_threads = summary["C2"].value
    issue_total = sum(issues.cell(r, 5).value or 0 for r in range(2, issues.max_row + 1))
    expected_lots = summary["E2"].value
    expected_wip = summary["G2"].value
    owner_lots = sum(owners.cell(r, 2).value or 0 for r in range(2, owners.max_row + 1))
    owner_wip = sum(owners.cell(r, 3).value or 0 for r in range(2, owners.max_row + 1))
    if issue_total != expected_threads:
        errors.append(f"Issue count total {issue_total} != Total Threads {expected_threads}")
    if owner_lots != expected_lots:
        errors.append(f"Owner lots total {owner_lots} != Current Onhold Lots {expected_lots}")
    if owner_wip != expected_wip:
        errors.append(f"Owner WIP total {owner_wip} != Current Onhold WIP {expected_wip}")

    for r in range(2, issues.max_row + 1):
        count = issues.cell(r, 5).value or 0
        shift_total = sum(issues.cell(r, c).value or 0 for c in range(6, 10))
        if count != shift_total:
            errors.append(f"Issue Contributors row {r}: Count {count} != shifts {shift_total}")

    if not isinstance(wb["Calendar Events"]["A2"].value, datetime):
        errors.append("Calendar Events A2 is not a typed Excel date")
    if not isinstance(wb["Feedback Evidence"]["D2"].value, datetime):
        errors.append("Feedback Evidence D2 is not a typed Excel date")

    compact_ranges = {}
    for ws in wb.worksheets:
        max_row = min(ws.max_row, 8)
        max_col = min(ws.max_column, 9)
        compact_ranges[ws.title] = [
            [ws.cell(row, col).value for col in range(1, max_col + 1)]
            for row in range(1, max_row + 1)
        ]

    result = {
        "file": str(path),
        "size_bytes": path.stat().st_size,
        "sheets": wb.sheetnames,
        "dimensions": {ws.title: ws.dimensions for ws in wb.worksheets},
        "table_counts": {ws.title: len(ws.tables) for ws in wb.worksheets},
        "validation_counts": {ws.title: len(ws.data_validations.dataValidation) for ws in wb.worksheets},
        "formula_cells": formulas,
        "formula_errors": formula_errors,
        "reconciliation": {
            "threads": {"summary": expected_threads, "issues": issue_total},
            "onhold_lots": {"summary": expected_lots, "owners": owner_lots},
            "onhold_wip": {"summary": expected_wip, "owners": owner_wip},
        },
        "compact_ranges": compact_ranges,
        "errors": errors,
    }
    if errors or formula_errors:
        raise RuntimeError(json.dumps(result, indent=2, default=str))
    return result


def main() -> None:
    output = build_workbook()
    previews = render_all_sheets(output)
    validation = validate_workbook(output)
    print(
        json.dumps(
            {
                "output": str(output),
                "previews": [str(path) for path in previews],
                "validation": validation,
            },
            indent=2,
            default=str,
        )
    )


if __name__ == "__main__":
    main()
